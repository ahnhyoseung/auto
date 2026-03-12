# ============================================================
# KOLAS 자동화 프로그램 v2.0
#
# [사이트]
#   https://www.knab.go.kr (한국인정기구 KOLAS 관리자 페이지)
#
# [전체 동작 순서]
#   1. 로그인
#   2. 평가계획서 목록 페이지에서 접수번호 검색
#   3. 목록에서 접수번호 링크 클릭 → 상세 페이지 진입
#      ※ 상세 페이지는 POST 방식이라 URL 직접 접근 불가 (400 에러)
#        반드시 링크 클릭으로 진입 후 그 상태 유지
#   4. 상세 페이지에서 아래 순서로 팝업 클릭하여 데이터 수집:
#      (1) 기관정보보기   → 기관명, 담당자명/연락처/이메일, 사업장주소
#      (2) 컨설팅정보보기 → 내부심사자(5년간심사자), 컨설팅담당자
#      (3) 인정분야보기   → 팝업 내 엑셀출력 클릭 → xls 다운로드
#      (4) 기관이력보기   → 정기검사/신규/재평가/확대 이력 중 최상단
#                          → 현장평가 처리완료 행의 평가반 성명 수집
#   5. template.xlsx 복사 후 '1안' 시트에 데이터 입력 저장
#
# [엑셀 셀 매핑 - '1안' 시트]
#   B3     = 기관명
#   D3     = 평가종류 (상세페이지 본문 "신청분류" 값: 신규/정기검사/재평가/확대)
#   B4     = 내부심사자 1번
#   C4     = 내부심사자 2번
#   B5     = 컨설팅 담당자 (없으면 빈칸)
#   B6~H6  = 직전 평가반 (최대 7명)
#   B7     = 평가 시작일
#   C7     = 평가 종료일
#   B14    = 담당자명
#   B15    = 담당자연락처
#   B16    = 담당자이메일
#   F14    = 사업장 주소
#   19행~  = 인정분야 데이터 (다운받은 xls 2행부터 복사, 1행은 헤더라 제외)
#
# [새 시트 자동 생성]
#   인정분야 xls에서 대분류코드(B열=인덱스1) + 중분류코드(D열=인덱스3) 조합으로
#   시트 자동 생성. 예: "01.017", "02.005"
#   xls에서 숫자는 float으로 읽히므로 int 변환 후 zfill 처리
#   예) 1.0 → int(1.0)=1 → zfill(2) → "01"
#       17.0 → int(17.0)=17 → zfill(3) → "017"
#
# [HTML 버튼 구조 - 상세 페이지]
#   기관정보보기:   <input type="button" value="기관정보보기"  onclick="PopupCompInfo('companyNo','01','02');" id="CompInfoPopUp">
#   컨설팅정보보기: <input type="button" value="컨설팅정보보기" onclick="PopupConsDoc('accreditNo');"       id="ConsDocPopUp">
#   인정분야보기:   <input type="button" value="인정분야보기"   onclick="PopupDocDetail111('accreditNo');"  id="PopupDocDetail">
#   기관이력보기:   <input type="button" value="기관이력보기"   onclick="OfcwbRcepPopUp('accreditNo');"     id="OfcwbRcepPopUpasb123">
#
# [기관정보보기 팝업 특이사항]
#   - 사업장 섹션 th: <th rowspan="6">사<br>업<br>장</th>
#     → row.text로는 "사업장" 감지 불가 → th.text.replace("\n","")로 감지
#   - 주소: <th>주소</th><td colspan="3">광주광역시 북구 첨단과기로 333</td>
#   - 법인/사업장 두 섹션 모두 "주소" th가 있음
#     → in_company_section 플래그로 사업장 섹션 진입 후 첫 번째 주소만 사용
#
# [컨설팅정보보기 팝업 특이사항]
#   - 내부심사자: <th>5년간심사자</th><td>KOLAS 선임평가사 김용오</td>
#     → "선임평가사" 또는 "평가사" 뒤의 한글 이름(2~4자) 추출
#   - 컨설팅 담당자: 테이블 4번째 열, "없음"이면 빈칸
#
# [인정분야 다운로드]
#   - 팝업 내 엑셀출력 버튼: <input class="TableBtn" onclick="javascript:XlsDownload();">
#   - 다운로드 파일명: "260309_133344508_ TestingRangeList.xls" (날짜+시간+고정텍스트)
#   - C:\Users\user\Downloads 에서 *TestingRangeList*.xls* 중 가장 최근 파일 사용
#
# [설치 명령]
#   pip install selenium openpyxl webdriver-manager xlrd
#
# [파일 구조]
#   C:\ahn\
#     ├── ahn.py          ← 이 스크립트
#     ├── template.xlsx   ← 양식 파일 ('1안' 시트 포함)
#     └── ahnhyoseung\    ← 결과 저장 폴더 (자동 생성)
# ============================================================

import re
import time
import shutil
import os
import glob
import openpyxl
import xlrd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# ── ★ 설정 ★ ──────────────────────────────────────────────
LOGIN_ID      = input("ID를 입력하세요:")        # 로그인 아이디
LOGIN_PW      = input('비밀번호를 입력하세요:')   # 로그인 비밀번호
BASE_URL      = "https://www.knab.go.kr"
LOGIN_URL     = "https://www.knab.go.kr/mgr/intr/lgn/LoginMngCpinsForm.do"
TEMPLATE_FILE = "template.xlsx"    # 양식 파일 (스크립트와 같은 폴더)
OUTPUT_DIR    = "ahnhyoseung"      # 결과 저장 폴더
DOWNLOAD_DIR  = r"C:\Users\user\Downloads"  # 인정분야 xls 다운로드 경로
# ──────────────────────────────────────────────────────────

# 기관이력 팝업에서 찾을 신청분류 키워드 (이 중 하나 포함된 행 선택)
HISTORY_KEYWORDS = ["정기검사", "신규", "재평가", "확대"]


# ════════════════════════════════════════════════════════════
# 공통 유틸
# ════════════════════════════════════════════════════════════

def setup_driver():
    """Chrome 드라이버 초기화. 팝업 차단 해제 옵션 포함."""
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-popup-blocking")  # 팝업창 허용
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.implicitly_wait(3)
    return driver


def wait_el(driver, by, selector, timeout=10):
    """특정 요소가 DOM에 나타날 때까지 최대 timeout초 대기."""
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, selector))
    )


def switch_to_popup(driver, main_window, timeout=8):
    """
    팝업창이 열릴 때까지 대기 후 팝업으로 포커스 전환.
    main_window: 메인 창 핸들 (driver.current_window_handle)
    반환: 팝업 창 핸들
    """
    WebDriverWait(driver, timeout).until(lambda d: len(d.window_handles) > 1)
    for handle in driver.window_handles:
        if handle != main_window:
            driver.switch_to.window(handle)
            return handle
    return None


def click_btn(driver, *selectors):
    """
    여러 셀렉터를 순서대로 시도하여 처음 찾은 요소 클릭.
    "//" 또는 "("로 시작하면 XPath, 아니면 CSS 셀렉터로 처리.
    성공하면 True, 모두 실패하면 False 반환.
    """
    for sel in selectors:
        try:
            if sel.startswith("//") or sel.startswith("("):
                el = driver.find_element(By.XPATH, sel)
            else:
                el = driver.find_element(By.CSS_SELECTOR, sel)
            el.click()
            return True
        except Exception:
            continue
    return False


# ════════════════════════════════════════════════════════════
# STEP 1: 로그인
# ════════════════════════════════════════════════════════════
def login(driver):
    """
    KOLAS 관리자 로그인.
    - 입력 필드: name="loginId", name="loginPw"
    - 로그인 버튼: <a href="javascript:Login();"> 형태라 JS 직접 호출
    """
    driver.get(LOGIN_URL)
    wait_el(driver, By.NAME, "loginId")
    driver.find_element(By.NAME, "loginId").clear()
    driver.find_element(By.NAME, "loginId").send_keys(LOGIN_ID)
    driver.find_element(By.NAME, "loginPw").clear()
    driver.find_element(By.NAME, "loginPw").send_keys(LOGIN_PW)
    try:
        driver.execute_script("Login();")  # JS 함수 직접 호출
    except Exception:
        click_btn(driver, "a[href='javascript:Login();']", "#loginBtn", "input[type='submit']")
    time.sleep(2)
    print("✅ 로그인 완료")


# ════════════════════════════════════════════════════════════
# STEP 2: 접수번호 검색 → 상세 페이지 진입
# ════════════════════════════════════════════════════════════
def collect_list(driver):
    """
    접수번호 입력받아 평가계획서 목록에서 검색 후 상세 페이지로 이동.

    ※ 중요: 상세 페이지(OfcwbEvlActplnStreRegistForm.do)는 POST 방식.
       URL만으로 재접근 시 400 에러 발생.
       반드시 목록에서 링크 클릭으로 진입해야 하며,
       이후 모든 팝업 작업은 이 페이지에서 수행해야 함.

    목록 링크 구조:
      <a href="#" onclick="fnReqActpl('companyNo','accreditNo','evlNo');">58916</a>
      → onclick에서 정규식으로 파라미터 추출 후 링크 직접 클릭
      → 새 창이 열리면 해당 창으로 전환

    반환: [{"접수번호": "58916", "href": "현재URL", "accreditNo": "58916"}, ...]
    """
    no_input = input("접수번호 입력: ").strip()
    accredit_nos = [no_input] if no_input else []

    items = []
    for no in accredit_nos:
        # 평가계획서 목록 페이지로 이동
        driver.get(BASE_URL + "/mgr/rcj/eva/OfcwbEvlActplnStreList.do")
        time.sleep(2)

        # 접수번호 입력칸에 검색어 입력 (name="searchAccreditNo")
        try:
            el = driver.find_element(By.NAME, "searchAccreditNo")
            el.clear()
            el.send_keys(no)
        except Exception:
            print(f"  ⚠️ 접수번호 입력칸 못 찾음")
            continue

        # 검색 버튼 클릭
        click_btn(driver, "input[value='검색']", "#searchBtn", ".btn-search")
        time.sleep(2)

        # 결과 목록에서 접수번호 링크 찾아 클릭
        try:
            rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
            for row in rows:
                try:
                    link = row.find_element(By.CSS_SELECTOR, "td a")
                    onclick = link.get_attribute("onclick") or ""
                    # fnReqActpl('7330','58916','14942') 형태에서 파라미터 추출
                    m = re.search(r"fnReqActpl\('(\d+)','(\d+)','(\d+)'\)", onclick)
                    if not m:
                        continue
                    accredit_no = m.group(2)  # 두 번째 파라미터 = accreditNo
                    main_window = driver.current_window_handle
                    link.click()
                    time.sleep(2)
                    # 새 창 열렸으면 전환
                    if len(driver.window_handles) > 1:
                        for handle in driver.window_handles:
                            if handle != main_window:
                                driver.switch_to.window(handle)
                                break
                    href = driver.current_url
                    items.append({"접수번호": no, "href": href, "accreditNo": accredit_no})
                    print(f"  ✅ 접수번호 {no} 찾음 → {href}")
                    break
                except Exception:
                    continue
        except Exception as e:
            print(f"  ⚠️ 접수번호 {no} 검색 오류: {e}")

    print(f"📋 총 {len(items)}개 접수 건 수집")
    return items


# ════════════════════════════════════════════════════════════
# STEP 3: 기관이력 팝업 → 직전 평가반 수집
# ════════════════════════════════════════════════════════════
def get_history_evaluators(driver, accredit_no):
    """
    기관이력보기 팝업에서 직전 평가반 성명 수집.
    버튼: <input type="button" value="기관이력보기" onclick="OfcwbRcepPopUp('accreditNo');">

    동작:
    1. 이력 목록에서 HISTORY_KEYWORDS(정기/신규/재평가/확대) 포함 최상단 행 클릭
    2. 이력 상세에서 진행단계="현장평가" + 처리상태="처리완료" 행 찾기
    3. 현장평가 상세에서 평가반 성명 수집
       - 구분 키워드: 반장, 평가사, 사보, 위원, 평가원
       - 테이블 구조: cells[0]=구분, cells[1]=소속, cells[2]=성명
    """
    evaluators = []
    main_window = driver.current_window_handle

    # 기관이력보기 버튼 클릭
    clicked = click_btn(
        driver,
        "//a[contains(text(),'기관이력보기')]",
        "//input[contains(@value,'기관이력보기')]",
        "//button[contains(text(),'기관이력보기')]"
    )
    if not clicked:
        # 버튼 못 찾으면 URL 직접 열기 시도
        url = f"{BASE_URL}/mgr/rcj/doj/OfcwbRcepPopUpList.do?accreditNo={accredit_no}"
        driver.execute_script(f"window.open('{url}', '_blank')")

    popup = switch_to_popup(driver, main_window)
    if not popup:
        print("  ⚠️ 기관이력 팝업 열리지 않음")
        return evaluators

    time.sleep(1.5)

    # 이력 목록에서 정기/신규/재평가/확대 포함 최상단 행 클릭
    target_href = None
    try:
        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
        for row in rows:
            cells = row.find_elements(By.TAG_NAME, "td")
            row_text = " ".join(c.text for c in cells)
            if any(kw in row_text for kw in HISTORY_KEYWORDS):
                try:
                    link = row.find_element(By.CSS_SELECTOR, "td a")
                    target_href = link.get_attribute("href")
                    print(f"  📎 이력 클릭: {link.text.strip()} ({row_text[:30]}...)")
                    break
                except Exception:
                    continue
    except Exception as e:
        print(f"  ⚠️ 기관이력 파싱 오류: {e}")

    driver.close()
    driver.switch_to.window(main_window)

    if not target_href:
        print("  ℹ️ 해당 이력 없음")
        return evaluators

    # 이력 상세 페이지 열기
    driver.execute_script(f"window.open('{target_href}', '_blank')")
    switch_to_popup(driver, main_window)
    time.sleep(1.5)

    # 현장평가 처리완료 행의 링크 찾기
    eval_href = None
    try:
        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
        for row in rows:
            cells = row.find_elements(By.TAG_NAME, "td")
            if len(cells) < 2:
                continue
            진행단계 = cells[0].text.strip()
            처리상태 = cells[1].text.strip()
            if "현장평가" in 진행단계 and "처리완료" in 처리상태:
                try:
                    eval_href = cells[0].find_element(By.TAG_NAME, "a").get_attribute("href")
                    print("  ✅ 현장평가 처리완료 확인")
                    break
                except Exception:
                    pass
    except Exception as e:
        print(f"  ⚠️ 이력 상세 파싱 오류: {e}")

    driver.close()
    driver.switch_to.window(main_window)

    if not eval_href:
        print("  ℹ️ 현장평가 처리완료 없음")
        return evaluators

    # 현장평가 상세에서 평가반 성명 수집
    driver.execute_script(f"window.open('{eval_href}', '_blank')")
    switch_to_popup(driver, main_window)
    time.sleep(1.5)

    evaluators = _extract_evaluators(driver)
    print(f"  👥 평가반: {evaluators}")

    driver.close()
    driver.switch_to.window(main_window)
    return evaluators


def _extract_evaluators(driver):
    """
    현장평가 상세 페이지에서 평가반 성명 추출.
    테이블 구조: cells[0]=구분, cells[1]=소속, cells[2]=성명
    구분에 반장/평가사/사보/위원/평가원 포함된 행만 수집.
    """
    names = []
    구분_keywords = ["반장", "평가사", "사보", "위원", "평가원"]
    try:
        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
        for row in rows:
            cells = row.find_elements(By.TAG_NAME, "td")
            if len(cells) >= 3:
                구분 = cells[0].text.strip()
                성명 = cells[2].text.strip()
                if 성명 and any(kw in 구분 for kw in 구분_keywords):
                    names.append(성명)
    except Exception as e:
        print(f"  ⚠️ 평가반 추출 오류: {e}")
    return names


# ════════════════════════════════════════════════════════════
# STEP 4: 컨설팅정보 팝업 → 내부심사자 + 컨설팅담당자 수집
# ════════════════════════════════════════════════════════════
def get_consulting_info(driver):
    """
    컨설팅정보보기 팝업에서 내부심사자와 컨설팅담당자 수집.
    버튼: <input type="button" value="컨설팅정보보기" onclick="PopupConsDoc('accreditNo');">

    내부심사자:
      <th>5년간심사자</th><td>KOLAS 선임평가사 김용오</td>
      → 정규식: "선임평가사" 또는 "평가사" 뒤의 한글 이름(2~4자) 추출
      예) "KOLAS 선임평가사 김용오" → ["김용오"]

    컨설팅담당자:
      테이블 행에서 cells[1]에 "컨설팅" 포함 시 cells[3]이 담당자명
      "없음"이면 수집하지 않음
    """
    result = {"내부심사": [], "컨설팅": []}
    main_window = driver.current_window_handle

    clicked = click_btn(
        driver,
        "//input[contains(@value,'컨설팅정보보기')]",
        "//a[contains(text(),'컨설팅정보보기')]"
    )
    if not clicked:
        print("  ⚠️ 컨설팅정보보기 버튼 없음")
        return result

    switch_to_popup(driver, main_window)
    time.sleep(1.5)

    # 5년간심사자 셀에서 이름 추출 ("평가사 이름" 패턴)
    try:
        cell_text = driver.find_element(
            By.XPATH, "//th[contains(text(),'5년간심사자')]/following-sibling::td"
        ).text
        names = re.findall(r'(?:선임)?평가사\s+([가-힣]{2,4})', cell_text)
        result["내부심사"] = names
    except Exception:
        pass

    # 컨설팅 담당자 수집 (4번째 열, "없음" 제외)
    try:
        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
        for row in rows:
            cells = row.find_elements(By.TAG_NAME, "td")
            if len(cells) >= 4 and "컨설팅" in cells[1].text:
                name = cells[3].text.strip()
                if name and name != "없음":
                    result["컨설팅"].append(name)
    except Exception:
        pass

    driver.close()
    driver.switch_to.window(main_window)
    return result


# ════════════════════════════════════════════════════════════
# STEP 5: 인정분야 팝업 → 엑셀 다운로드
# ════════════════════════════════════════════════════════════
def get_accredit_range(driver):
    """
    인정분야보기 팝업에서 엑셀 다운로드 후 파일 경로 반환.
    버튼: <input type="button" value="인정분야보기" onclick="PopupDocDetail111('accreditNo');">

    팝업 내 엑셀출력:
      <input class="TableBtn" onclick="javascript:XlsDownload();">
      → JS 함수 직접 호출

    다운로드 파일명: "260309_133344508_ TestingRangeList.xls" (날짜+시간+고정텍스트)
    → DOWNLOAD_DIR에서 *TestingRangeList*.xls* 중 가장 최근 파일(mtime 기준) 사용

    반환: 다운로드된 xls 파일 경로 (실패 시 None)
    """
    main_window = driver.current_window_handle

    clicked = click_btn(
        driver,
        "//input[contains(@value,'인정분야보기')]",
        "//a[contains(text(),'인정분야보기')]"
    )
    if not clicked:
        print("  ⚠️ 인정분야보기 버튼 없음")
        return None

    switch_to_popup(driver, main_window)
    time.sleep(2)

    # 엑셀출력 JS 함수 호출
    try:
        driver.execute_script("XlsDownload();")
        time.sleep(3)  # 다운로드 완료 대기
    except Exception as e:
        print(f"  ⚠️ 엑셀출력 오류: {e}")
        driver.close()
        driver.switch_to.window(main_window)
        return None

    # Downloads 폴더에서 가장 최근 TestingRangeList 파일 찾기
    files = glob.glob(os.path.join(DOWNLOAD_DIR, "*TestingRangeList*.xls*"))
    if not files:
        print("  ⚠️ 다운로드 파일 못 찾음")
        driver.close()
        driver.switch_to.window(main_window)
        return None

    latest = max(files, key=os.path.getmtime)
    print(f"  📥 다운로드: {os.path.basename(latest)}")

    driver.close()
    driver.switch_to.window(main_window)
    return latest


# ════════════════════════════════════════════════════════════
# STEP 6: 기관정보 팝업 → 기관명, 담당자정보, 사업장주소 수집
# ════════════════════════════════════════════════════════════
def get_agency_info(driver):
    """
    기관정보보기 팝업에서 기관명, 담당자 정보, 사업장 주소 수집.
    버튼: <input type="button" value="기관정보보기" onclick="PopupCompInfo('companyNo','01','02');" id="CompInfoPopUp">

    팝업 HTML 특이사항:
    1. 사업장 섹션 감지:
       <th rowspan="6">사<br>업<br>장</th> 형태로 세로 병합
       → th.text = "사\n업\n장" → replace("\n","") 후 "사업장" 감지
       → in_company_section 플래그를 True로 설정

    2. 주소 수집:
       <th>주소</th><td colspan="3">광주광역시 북구 첨단과기로 333</td>
       → th.find_element(XPATH, "following-sibling::td[1]")로 td 가져옴
       → 법인/사업장 두 곳에 "주소" th가 있으므로
         in_company_section=True 이후 첫 번째 주소만 수집

    3. 담당자 정보는 팝업 하단에 위치 (스크롤 없이도 DOM에 있음)
    """
    result = {"기관명": "", "담당자명": "", "담당자연락처": "", "담당자이메일": "", "주소": ""}
    main_window = driver.current_window_handle

    clicked = click_btn(
        driver,
        "//input[contains(@value,'기관정보보기')]",
        "//a[contains(text(),'기관정보보기')]"
    )
    if not clicked:
        print("  ⚠️ 기관정보보기 버튼 없음")
        return result

    switch_to_popup(driver, main_window)
    time.sleep(2)

    try:
        rows = driver.find_elements(By.CSS_SELECTOR, "table tr")
        in_company_section = False  # 사업장 섹션 진입 여부 플래그

        for row in rows:
            ths = row.find_elements(By.TAG_NAME, "th")
            tds = row.find_elements(By.TAG_NAME, "td")

            # 사업장 rowspan th 감지: "사\n업\n장" → replace로 "사업장"
            for th in ths:
                th_text = th.text.replace("\n", "").replace(" ", "")
                if "사업장" in th_text:
                    in_company_section = True

            for th in ths:
                label = th.text.replace("\n", "").strip()
                # th 바로 다음 td 가져오기 (following-sibling)
                try:
                    td = th.find_element(By.XPATH, "following-sibling::td[1]")
                    val = td.text.strip()
                except Exception:
                    val = tds[0].text.strip() if tds else ""

                if "기관명" in label and not result["기관명"] and "영문" not in label:
                    result["기관명"] = val
                elif "담당자명" in label:
                    result["담당자명"] = val
                elif "담당자연락처" in label or ("연락처" in label and "담당" in label):
                    result["담당자연락처"] = val
                elif "담당자이메일" in label or ("이메일" in label and "담당" in label):
                    result["담당자이메일"] = val
                elif label == "주소" and in_company_section and not result["주소"]:
                    # 사업장 섹션의 첫 번째 "주소"만 수집 (법인 주소 제외)
                    result["주소"] = val

    except Exception as e:
        print(f"  ⚠️ 기관정보 파싱 오류: {e}")

    driver.close()
    driver.switch_to.window(main_window)
    return result


# ════════════════════════════════════════════════════════════
# STEP 7: 엑셀 템플릿 복사 후 데이터 입력 저장
# ════════════════════════════════════════════════════════════
def fill_excel(data):
    """
    template.xlsx를 복사하여 '1안' 시트에 데이터 입력 후 저장.

    인정분야 데이터:
    - xlrd로 xls 읽기 (openpyxl은 .xls 미지원, xlrd 사용)
    - 1행(인덱스0)은 헤더 → 2행(인덱스1)부터 '1안' 시트 19행에 붙여넣기
    - 대분류코드(B열=인덱스1) + 중분류코드(D열=인덱스3) 조합으로 새 시트 생성
      xls 숫자는 float으로 읽힘 → int 변환 후 zfill 처리
      예) 1.0 → "01", 17.0 → "017" → 시트명 "01.017"

    파일명: {기관명}_{평가종류}_평가계획서(안).xlsx
    저장위치: OUTPUT_DIR (ahnhyoseung/)
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    기관명   = data.get("기관명", "unknown")
    평가종류 = data.get("평가종류", "")
    # 파일명 특수문자 제거
    safe_name = re.sub(r'[\\/*?:"<>|]', "_", 기관명)
    out_path  = os.path.join(OUTPUT_DIR, f"{safe_name}_{평가종류}_평가계획서(안).xlsx")

    # 템플릿 복사 후 열기
    shutil.copy2(TEMPLATE_FILE, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb["1안"]  # '1안' 시트에 입력

    # 기본 정보
    ws["B3"] = 기관명
    ws["D3"] = 평가종류

    # 내부심사자 (최대 2명)
    심사자 = data.get("내부심사", [])
    ws["B4"] = 심사자[0] if len(심사자) > 0 else ""
    ws["C4"] = 심사자[1] if len(심사자) > 1 else ""

    # 컨설팅 담당자
    컨설팅 = data.get("컨설팅", [])
    ws["B5"] = 컨설팅[0] if 컨설팅 else ""

    # 직전 평가반 (B6~H6, 최대 7명)
    cols  = ["B", "C", "D", "E", "F", "G", "H"]
    평가반 = data.get("직전평가반", [])
    for i, col in enumerate(cols):
        ws[f"{col}6"] = 평가반[i] if i < len(평가반) else ""

    # 평가일
    ws["B7"] = data.get("평가시작일", "")
    ws["C7"] = data.get("평가종료일", "")

    # 기관 담당자 정보
    ws["B14"] = data.get("담당자명", "")
    ws["B15"] = data.get("담당자연락처", "")
    ws["B16"] = data.get("담당자이메일", "")
    ws["F14"] = data.get("주소", "")  # 사업장 주소

    # 인정분야 데이터 붙여넣기 + 대분류.중분류 시트 생성
    range_file = data.get("인정분야파일", None)
    if range_file and os.path.exists(range_file):
        try:
            # xlrd로 .xls 파일 읽기 (openpyxl은 .xls 미지원)
            wb_range = xlrd.open_workbook(range_file)
            ws_range = wb_range.sheet_by_index(0)

            # '1안' 시트 19행부터 붙여넣기 (xls 2행=인덱스1부터, 1행은 헤더)
            start_row = 19
            for row_idx in range(1, ws_range.nrows):
                for col_idx in range(ws_range.ncols):
                    val = ws_range.cell_value(row_idx, col_idx)
                    ws.cell(row=start_row, column=col_idx + 1, value=val)
                start_row += 1
            print(f"  📋 인정분야 {start_row - 19}행 복사 완료")

            # 대분류코드(B열=인덱스1) + 중분류코드(D열=인덱스3) 조합으로 시트 생성
            sheet_names = set()
            for row_idx in range(1, ws_range.nrows):
                대분류_val = ws_range.cell_value(row_idx, 1)  # B열 (인덱스1)
                중분류_val = ws_range.cell_value(row_idx, 3)  # D열 (인덱스3)
                # xls에서 숫자는 float으로 읽힘 → int 변환 후 zfill
                대분류 = str(int(float(대분류_val))).zfill(2) if 대분류_val else ""
                중분류 = str(int(float(중분류_val))).zfill(3) if 중분류_val else ""
                if 대분류 and 중분류:
                    sheet_names.add(f"{대분류}.{중분류}")

            for sheet_name in sorted(sheet_names):
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(title=sheet_name)
                    print(f"  📄 시트 생성: {sheet_name}")

        except Exception as e:
            print(f"  ⚠️ 인정분야 복사 오류: {e}")

    wb.save(out_path)
    print(f"  💾 저장: {out_path}")
    return out_path


# ════════════════════════════════════════════════════════════
# 메인
# ════════════════════════════════════════════════════════════
def main():
    print("=" * 60)
    print("  KOLAS 자동화 프로그램 v2.0")
    print("=" * 60)

    if not os.path.exists(TEMPLATE_FILE):
        print(f"❌ 템플릿 파일 없음: {TEMPLATE_FILE}")
        return

    driver = setup_driver()
    results = []

    try:
        login(driver)
        items = collect_list(driver)

        if not items:
            print("❌ 수집된 항목 없음.")
            return

        for idx, item in enumerate(items, 1):
            접수번호 = item["접수번호"]
            print(f"\n[{idx}/{len(items)}] 접수번호: {접수번호}")

            try:
                accredit = item["accreditNo"]

                # 평가종류: 상세페이지 본문 "신청분류" 셀에서 수집
                # (신규/정기검사/재평가/확대 중 하나)
                평가종류 = ""
                try:
                    rows = driver.find_elements(By.CSS_SELECTOR, "table tr")
                    for row in rows:
                        ths = row.find_elements(By.TAG_NAME, "th")
                        tds = row.find_elements(By.TAG_NAME, "td")
                        for i, th in enumerate(ths):
                            if "신청분류" in th.text.strip():
                                평가종류 = tds[i].text.strip() if i < len(tds) else ""
                                break
                        if 평가종류:
                            break
                except Exception:
                    pass

                # 팝업 수집 순서 (이 순서 중요!)
                # 상세 페이지는 POST 방식이라 팝업 닫고 URL 재접근 불가
                # 따라서 모든 팝업을 한 번의 상세페이지 진입 상태에서 처리

                # 1. 기관정보보기 팝업 → 기관명, 담당자정보, 사업장주소
                agency = get_agency_info(driver)

                # 2. 컨설팅정보보기 팝업 → 내부심사자, 컨설팅담당자
                consulting = get_consulting_info(driver)

                # 3. 인정분야보기 팝업 → 엑셀 다운로드
                range_file = get_accredit_range(driver)

                # 4. 기관이력보기 팝업 → 직전 평가반 (마지막 수행)
                evaluators = get_history_evaluators(driver, accredit)

                data = {
                    "기관명":       agency.get("기관명", ""),
                    "평가종류":     평가종류,
                    "내부심사":     consulting.get("내부심사", []),
                    "컨설팅":       consulting.get("컨설팅", []),
                    "직전평가반":   evaluators,
                    "평가시작일":   "",   # 현재 미사용 (추후 구현 가능)
                    "평가종료일":   "",   # 현재 미사용 (추후 구현 가능)
                    "담당자명":     agency.get("담당자명", ""),
                    "담당자연락처": agency.get("담당자연락처", ""),
                    "담당자이메일": agency.get("담당자이메일", ""),
                    "주소":         agency.get("주소", ""),
                    "인정분야파일": range_file,
                }

                print(f"  기관명:     {data['기관명']}")
                print(f"  평가종류:   {data['평가종류']}")
                print(f"  내부심사:   {data['내부심사']}")
                print(f"  컨설팅:     {data['컨설팅']}")
                print(f"  직전평가반: {data['직전평가반']}")
                print(f"  담당자:     {data['담당자명']} / {data['담당자연락처']}")

                out = fill_excel(data)
                results.append({"접수번호": 접수번호, "기관명": data["기관명"], "파일": out})

            except Exception as e:
                print(f"  ❌ 처리 중 오류: {e}")
                import traceback
                traceback.print_exc()
                # 오류 발생 시 메인 창으로 복귀 후 다음 접수번호 처리 계속
                try:
                    driver.switch_to.window(driver.window_handles[0])
                except Exception:
                    pass
                continue

    except Exception as e:
        print(f"\n❌ 치명적 오류: {e}")
        import traceback
        traceback.print_exc()

    finally:
        driver.quit()

    print("\n" + "=" * 60)
    print(f"  완료! 총 {len(results)}개 파일 → '{OUTPUT_DIR}/' 폴더")
    print("=" * 60)
    for r in results:
        print(f"  [{r['접수번호']}] {r['기관명']}")


if __name__ == "__main__":
    main()