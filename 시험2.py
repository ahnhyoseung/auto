# ============================================================
# KOLAS 자동화 프로그램 v2.0
#
# [사이트]
#   https://www.knab.go.kr (한국인정기구 KOLAS 관리자 페이지)
#
# [전체 동작 순서]
#   1. 로그인
#   2. 평가계획서 목록 페이지에서 접수번호 검색
#   3. 목록에서 접수번호 링크 클릭 → 상세 페이지 진입
#      ※ 상세 페이지는 POST 방식이라 URL 직접 접근 불가 (400 에러)
#        반드시 링크 클릭으로 진입 후 그 상태 유지
#   4. 상세 페이지에서 아래 순서로 팝업 클릭하여 데이터 수집:
#      (1) 기관정보보기   → 기관명, 담당자명/연락처/이메일, 사업장주소
#      (2) 컨설팅정보보기 → 내부심사자(5년간심사자), 컨설팅담당자
#      (3) 인정분야보기   → 팝업 내 엑셀출력 클릭 → xls 다운로드
#      (4) 기관이력보기   → 정기검사/신규/재평가/확대 이력 중 최상단
#                          → 현장평가 처리완료 행의 평가반 성명 수집
#   5. template.xlsx 복사 후 '1안' 시트에 데이터 입력 저장
#
# [엑셀 셀 매핑 - '1안' 시트]
#   B3     = 기관명
#   D3     = 평가종류 (상세페이지 본문 "신청분류" 값: 신규/정기검사/재평가/확대)
#   B4     = 내부심사자 1번
#   C4     = 내부심사자 2번
#   B5     = 컨설팅 담당자 (없으면 빈칸)
#   B6~H6  = 직전 평가반 (최대 7명)
#   B7     = 평가 시작일
#   C7     = 평가 종료일
#   B14    = 담당자명
#   B15    = 담당자연락처
#   B16    = 담당자이메일
#   F14    = 사업장 주소
#   19행~  = 인정분야 데이터 (다운받은 xls 2행부터 복사, 1행은 헤더라 제외)
#
# [새 시트 자동 생성]
#   인정분야 xls에서 대분류코드(B열=인덱스1) + 중분류코드(D열=인덱스3) 조합으로
#   시트 자동 생성. 예: "01.017", "02.005"
#   xls에서 숫자는 float으로 읽히므로 int 변환 후 zfill 처리
#   예) 1.0 → int(1.0)=1 → zfill(2) → "01"
#       17.0 → int(17.0)=17 → zfill(3) → "017"
#
# [HTML 버튼 구조 - 상세 페이지]
#   기관정보보기:   <input type="button" value="기관정보보기"  onclick="PopupCompInfo('companyNo','01','02');" id="CompInfoPopUp">
#   컨설팅정보보기: <input type="button" value="컨설팅정보보기" onclick="PopupConsDoc('accreditNo');"       id="ConsDocPopUp">
#   인정분야보기:   <input type="button" value="인정분야보기"   onclick="PopupDocDetail111('accreditNo');"  id="PopupDocDetail">
#   기관이력보기:   <input type="button" value="기관이력보기"   onclick="OfcwbRcepPopUp('accreditNo');"     id="OfcwbRcepPopUpasb123">
#
# [기관정보보기 팝업 특이사항]
#   - 사업장 섹션 th: <th rowspan="6">사<br>업<br>장</th>
#     → row.text로는 "사업장" 감지 불가 → th.text.replace("\n","")로 감지
#   - 주소: <th>주소</th><td colspan="3">광주광역시 북구 첨단과기로 333</td>
#   - 법인/사업장 두 섹션 모두 "주소" th가 있음
#     → in_company_section 플래그로 사업장 섹션 진입 후 첫 번째 주소만 사용
#
# [컨설팅정보보기 팝업 특이사항]
#   - 내부심사자: <th>5년간심사자</th><td>KOLAS 선임평가사 김용오</td>
#     → "선임평가사" 또는 "평가사" 뒤의 한글 이름(2~4자) 추출
#   - 컨설팅 담당자: 테이블 4번째 열, "없음"이면 빈칸
#
# [인정분야 다운로드]
#   - 팝업 내 엑셀출력 버튼: <input class="TableBtn" onclick="javascript:XlsDownload();">
#   - 다운로드 파일명: "260309_133344508_ TestingRangeList.xls" (날짜+시간+고정텍스트)
#   - C:\Users\user\Downloads 에서 *TestingRangeList*.xls* 중 가장 최근 파일 사용
#
# [설치 명령]
#   pip install selenium openpyxl webdriver-manager xlrd
#
# [파일 구조]
#   C:\ahn\
#     ├── ahn.py          ← 이 스크립트
#     ├── template.xlsx   ← 양식 파일 ('1안' 시트 포함)
#     └── ahnhyoseung\    ← 결과 저장 폴더 (자동 생성)
# ============================================================

import re
import sys
import os

# exe로 배포 시 실행 파일 기준 경로 설정
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

import time
from dotenv import load_dotenv
load_dotenv()
import shutil
import glob
import openpyxl
import xlrd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# ── ★ 설정 ★ ──────────────────────────────────────────────
LOGIN_ID      = os.getenv("KOLAS_ID") or input("ID를 입력하세요: ")
LOGIN_PW      = os.getenv("KOLAS_PW") or input("비밀번호를 입력하세요: ")
BASE_URL      = "https://www.knab.go.kr"
LOGIN_URL     = "https://www.knab.go.kr/mgr/intr/lgn/LoginMngCpinsForm.do"
TEMPLATE_FILE = os.path.join(BASE_DIR, "template.xlsx")  # 양식 파일 (실행파일과 같은 폴더)
OUTPUT_DIR    = os.path.join(BASE_DIR, "ahnhyoseung")          # 결과 저장 폴더
DOWNLOAD_DIR  = os.getenv("DOWNLOAD_DIR") or os.path.join(os.path.expanduser("~"), "Downloads")
# ──────────────────────────────────────────────────────────

# 기관이력 팝업에서 찾을 신청분류 키워드 (이 중 하나 포함된 행 선택)
HISTORY_KEYWORDS = ["정기검사", "신규", "재평가", "확대"]


# ════════════════════════════════════════════════════════════
# 공통 유틸
# ════════════════════════════════════════════════════════════

def setup_driver():
    """Chrome 드라이버 초기화. 팝업 차단 해제 옵션 포함."""
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-popup-blocking")  # 팝업창 허용
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")  # 자동화 감지 우회
    options.add_experimental_option("excludeSwitches", ["enable-automation"])  # 자동화 배너 제거
    options.add_experimental_option("useAutomationExtension", False)  # 자동화 확장 비활성화
    # 비밀번호 저장 팝업 비활성화
    prefs = {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False
    }
    options.add_experimental_option("prefs", prefs)
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.implicitly_wait(3)
    return driver


def wait_el(driver, by, selector, timeout=10):
    """특정 요소가 DOM에 나타날 때까지 최대 timeout초 대기."""
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, selector))
    )


def switch_to_popup(driver, main_window, timeout=8):
    """
    팝업창이 열릴 때까지 대기 후 팝업으로 포커스 전환.
    main_window: 메인 창 핸들 (driver.current_window_handle)
    반환: 팝업 창 핸들
    """
    WebDriverWait(driver, timeout).until(lambda d: len(d.window_handles) > 1)
    for handle in driver.window_handles:
        if handle != main_window:
            driver.switch_to.window(handle)
            return handle
    return None


def click_btn(driver, *selectors):
    """
    여러 셀렉터를 순서대로 시도하여 처음 찾은 요소 클릭.
    "//" 또는 "("로 시작하면 XPath, 아니면 CSS 셀렉터로 처리.
    성공하면 True, 모두 실패하면 False 반환.
    """
    for sel in selectors:
        try:
            if sel.startswith("//") or sel.startswith("("):
                el = driver.find_element(By.XPATH, sel)
            else:
                el = driver.find_element(By.CSS_SELECTOR, sel)
            el.click()
            return True
        except Exception:
            continue
    return False


# ════════════════════════════════════════════════════════════
# STEP 1: 로그인
# ════════════════════════════════════════════════════════════
def login(driver):
    """
    KOLAS 관리자 로그인.
    - 입력 필드: name="loginId", name="loginPw"
    - 로그인 버튼: <a href="javascript:Login();"> 형태라 JS 직접 호출
    """
    driver.get(LOGIN_URL)
    wait_el(driver, By.NAME, "loginId")
    driver.find_element(By.NAME, "loginId").clear()
    driver.find_element(By.NAME, "loginId").send_keys(LOGIN_ID)
    driver.find_element(By.NAME, "loginPw").clear()
    driver.find_element(By.NAME, "loginPw").send_keys(LOGIN_PW)
    try:
        driver.execute_script("Login();")  # JS 함수 직접 호출
    except Exception:
        click_btn(driver, "a[href='javascript:Login();']", "#loginBtn", "input[type='submit']")
    time.sleep(2)
    print("✅ 로그인 완료")


# ════════════════════════════════════════════════════════════
# STEP 2: 접수번호 검색 → 상세 페이지 진입
# ════════════════════════════════════════════════════════════
def collect_list(driver):
    """
    접수번호 입력받아 평가계획서 목록에서 검색 후 상세 페이지로 이동.

    ※ 중요: 상세 페이지(OfcwbEvlActplnStreRegistForm.do)는 POST 방식.
       URL만으로 재접근 시 400 에러 발생.
       반드시 목록에서 링크 클릭으로 진입해야 하며,
       이후 모든 팝업 작업은 이 페이지에서 수행해야 함.

    목록 링크 구조:
      <a href="#" onclick="fnReqActpl('companyNo','accreditNo','evlNo');">58916</a>
      → onclick에서 정규식으로 파라미터 추출 후 링크 직접 클릭
      → 새 창이 열리면 해당 창으로 전환

    반환: [{"접수번호": "58916", "href": "현재URL", "accreditNo": "58916"}, ...]
    """
    no_input = input("접수번호 입력: ").strip()
    accredit_nos = [no_input] if no_input else []

    items = []
    for no in accredit_nos:
        # 평가계획서 목록 페이지로 이동
        driver.get(BASE_URL + "/mgr/rcj/eva/OfcwbEvlActplnStreList.do")
        time.sleep(2)

        # 접수번호 입력칸에 검색어 입력 (name="searchAccreditNo")
        try:
            el = driver.find_element(By.NAME, "searchAccreditNo")
            el.clear()
            el.send_keys(no)
        except Exception:
            print(f"  ⚠️ 접수번호 입력칸 못 찾음")
            continue

        # 검색 버튼 클릭
        click_btn(driver, "input[value='검색']", "#searchBtn", ".btn-search")
        time.sleep(2)

        # 결과 목록에서 접수번호 링크 찾아 클릭
        try:
            rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
            for row in rows:
                try:
                    link = row.find_element(By.CSS_SELECTOR, "td a")
                    onclick = link.get_attribute("onclick") or ""
                    # fnReqActpl('7330','58916','14942') 형태에서 파라미터 추출
                    m = re.search(r"fnReqActpl\('(\d+)','(\d+)','(\d+)'\)", onclick)
                    if not m:
                        continue
                    accredit_no = m.group(2)  # 두 번째 파라미터 = accreditNo
                    main_window = driver.current_window_handle
                    link.click()
                    time.sleep(2)
                    # 새 창 열렸으면 전환
                    if len(driver.window_handles) > 1:
                        for handle in driver.window_handles:
                            if handle != main_window:
                                driver.switch_to.window(handle)
                                break
                    href = driver.current_url
                    items.append({"접수번호": no, "href": href, "accreditNo": accredit_no})
                    print(f"  ✅ 접수번호 {no} 찾음 → {href}")
                    break
                except Exception:
                    continue
        except Exception as e:
            print(f"  ⚠️ 접수번호 {no} 검색 오류: {e}")

    print(f"📋 총 {len(items)}개 접수 건 수집")
    return items


# ════════════════════════════════════════════════════════════
# STEP 3: 기관이력 팝업 → 직전 평가반 수집
# ════════════════════════════════════════════════════════════
def get_history_evaluators(driver, accredit_no):
    """
    기관이력보기 팝업 → 이력 행 클릭 → 현장평가 행 평가계획서 클릭 → 평가반 수집
    팝업을 닫지 않고 연속 클릭으로 처리 (POST 방식 페이지라 URL 직접 접근 불가)
    """
    evaluators = []
    main_window = driver.current_window_handle

    # 1. 기관이력보기 팝업 열기
    clicked = click_btn(
        driver,
        "//a[contains(text(),'기관이력보기')]",
        "//input[contains(@value,'기관이력보기')]",
        "//button[contains(text(),'기관이력보기')]"
    )
    if not clicked:
        url = f"{BASE_URL}/mgr/rcj/doj/OfcwbRcepPopUpList.do?accreditNo={accredit_no}"
        driver.execute_script(f"window.open('{url}', '_blank')")

    popup1 = switch_to_popup(driver, main_window)
    if not popup1:
        print("  ⚠️ 기관이력 팝업 열리지 않음")
        return evaluators
    time.sleep(1.5)

    # 2. 이력 목록에서 HISTORY_KEYWORDS 포함 최상단 행 접수번호 클릭
    # HistoryInfo()는 새 창을 열음
    try:
        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
        for row in rows:
            cells = row.find_elements(By.TAG_NAME, "td")
            row_text = " ".join(c.text for c in cells)
            if any(kw in row_text for kw in HISTORY_KEYWORDS):
                try:
                    link = row.find_element(By.CSS_SELECTOR, "td a")
                    rcep_no = link.text.strip()
                    print(f"  📎 이력 클릭: {rcep_no}")
                    link.click()
                    time.sleep(3)
                    # 팝업 닫고 메인 창으로 전환
                    driver.close()  # 팝업 닫기
                    driver.switch_to.window(main_window)
                    print(f"  🔍 메인 창 URL: {driver.current_url}")
                    break
                except Exception as ex:
                    print(f"  ⚠️ 이력 행 클릭 오류: {ex}")
                    continue
    except Exception as e:
        print(f"  ⚠️ 기관이력 파싱 오류: {e}")
        for h in list(driver.window_handles):
            if h != main_window:
                driver.switch_to.window(h)
                driver.close()
        driver.switch_to.window(main_window)
        return evaluators

    time.sleep(1)

    # 3. 새 창에서 goProcSttus 링크 찾아 클릭 → 또 새 창으로 열림
    try:
        plan_link = driver.find_element(By.XPATH, "//a[contains(@href,'goProcSttus') and contains(text(),'평가계획서')]")
        href = plan_link.get_attribute("href")
        print(f"  ✅ 평가계획서 클릭")
        before2 = set(driver.window_handles)
        driver.execute_script(href.replace("javascript:", ""))
        time.sleep(2)
        after2 = set(driver.window_handles)
        new_win = after2 - before2
        print(f"  🔍 평가계획서 후 창 수: {len(driver.window_handles)}, 새창: {len(new_win)}")
        if new_win:
            driver.switch_to.window(list(new_win)[0])
        print(f"  🔍 평가계획서 창 URL: {driver.current_url}")
        evaluators = _extract_evaluators(driver)
        print(f"  👥 직전평가반: {evaluators}")
        for h in list(driver.window_handles):
            if h != main_window:
                driver.switch_to.window(h)
                driver.close()
        driver.switch_to.window(main_window)
        return evaluators
    except Exception as e:
        print(f"  ⚠️ 평가계획서 클릭 오류: {e}")

    # 열린 창들 정리
    for h in list(driver.window_handles):
        if h != main_window:
            driver.switch_to.window(h)
            driver.close()
    driver.switch_to.window(main_window)

    print("  ℹ️ 현장평가 평가계획서 없음")
    return evaluators


def _extract_evaluators(driver):
    """
    평가계획서 페이지에서 평가반 성명 추출.
    구조: 구분(1열) | 중분류(2열) | 성명(3열) | 소속(4열) ...
    구분에 반장/평가사/사보/위원/평가원 포함된 행의 성명(3열) 수집
    """
    names = []
    구분_keywords = ["반장", "평가사", "사보", "위원", "평가원"]
    try:
        rows = driver.find_elements(By.CSS_SELECTOR, "table tr")
        for row in rows:
            cells = row.find_elements(By.TAG_NAME, "td")
            if len(cells) >= 3:
                구분 = cells[0].text.strip()
                성명 = cells[2].text.strip()
                if 성명 and any(kw in 구분 for kw in 구분_keywords):
                    names.append(성명)
    except Exception as e:
        print(f"  ⚠️ 평가반 추출 오류: {e}")
    return names


# ════════════════════════════════════════════════════════════
# STEP 4: 컨설팅정보 팝업 → 내부심사자 + 컨설팅담당자 수집
# ════════════════════════════════════════════════════════════
def get_consulting_info(driver):
    """
    컨설팅정보보기 팝업에서 내부심사자와 컨설팅담당자 수집.
    버튼: <input type="button" value="컨설팅정보보기" onclick="PopupConsDoc('accreditNo');">

    내부심사자:
      <th>5년간심사자</th><td>KOLAS 선임평가사 김용오</td>
      → 정규식: "선임평가사" 또는 "평가사" 뒤의 한글 이름(2~4자) 추출
      예) "KOLAS 선임평가사 김용오" → ["김용오"]

    컨설팅담당자:
      테이블 행에서 cells[1]에 "컨설팅" 포함 시 cells[3]이 담당자명
      "없음"이면 수집하지 않음
    """
    result = {"내부심사": [], "컨설팅": []}
    main_window = driver.current_window_handle

    clicked = click_btn(
        driver,
        "//input[contains(@value,'컨설팅정보보기')]",
        "//a[contains(text(),'컨설팅정보보기')]"
    )
    if not clicked:
        print("  ⚠️ 컨설팅정보보기 버튼 없음")
        return result

    switch_to_popup(driver, main_window)
    time.sleep(1.5)

    # 5년간심사자 셀에서 이름 추출
    # 패턴1: "심사반장:홍길동", "심사원:홍길동" 형태
    # 패턴2: "KOLAS 선임평가사 홍길동" 형태
    try:
        cell_text = driver.find_element(
            By.XPATH, "//th[contains(text(),'5년간심사자')]/following-sibling::td"
        ).text
        # 패턴1: 심사반장/심사원: 뒤 이름
        names = re.findall(r'(?:심사반장|심사원):([가-힣]{2,4})', cell_text)
        if not names:
            # 패턴2: "KOLAS 선임평가사 이름" 또는 "이름(KOLAS...)" 둘 다 처리
            p2 = re.findall(r'(?:선임)?평가사\s+([가-힣]{2,4})', cell_text)
            p3 = re.findall(r'([가-힣]{2,4})\s*\(KOLAS', cell_text)
            names = p2 + p3
        # 중복 제거 (순서 유지)
        seen = set()
        unique_names = []
        for n in names:
            if n not in seen:
                seen.add(n)
                unique_names.append(n)
        result["내부심사"] = unique_names
    except Exception:
        pass

    # 컨설팅 담당자 수집
    # tbody 각 행의 4번째 td(인덱스3)가 담당자명
    try:
        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
        for row in rows:
            cells = row.find_elements(By.TAG_NAME, "td")
            if len(cells) >= 4:
                name = cells[3].text.strip()
                if name and name != "없음":
                    result["컨설팅"].append(name)
    except Exception:
        pass

    driver.close()
    driver.switch_to.window(main_window)
    return result


# ════════════════════════════════════════════════════════════
# STEP 5: 인정분야 팝업 → 엑셀 다운로드
# ════════════════════════════════════════════════════════════
def get_accredit_range(driver):
    """
    인정분야보기 팝업에서 엑셀 다운로드 후 파일 경로 반환.
    버튼: <input type="button" value="인정분야보기" onclick="PopupDocDetail111('accreditNo');">

    팝업 내 엑셀출력:
      <input class="TableBtn" onclick="javascript:XlsDownload();">
      → JS 함수 직접 호출

    다운로드 파일명: "260309_133344508_ TestingRangeList.xls" (날짜+시간+고정텍스트)
    → DOWNLOAD_DIR에서 *TestingRangeList*.xls* 중 가장 최근 파일(mtime 기준) 사용

    반환: 다운로드된 xls 파일 경로 (실패 시 None)
    """
    main_window = driver.current_window_handle

    clicked = click_btn(
        driver,
        "//input[contains(@value,'인정분야보기')]",
        "//a[contains(text(),'인정분야보기')]"
    )
    if not clicked:
        print("  ⚠️ 인정분야보기 버튼 없음")
        return None

    switch_to_popup(driver, main_window)
    time.sleep(2)

    # 엑셀출력 JS 함수 호출
    try:
        driver.execute_script("XlsDownload();")
        time.sleep(3)  # 다운로드 완료 대기
    except Exception as e:
        print(f"  ⚠️ 엑셀출력 오류: {e}")
        driver.close()
        driver.switch_to.window(main_window)
        return None

    # Downloads 폴더에서 가장 최근 TestingRangeList 파일 찾기
    files = glob.glob(os.path.join(DOWNLOAD_DIR, "*TestingRangeList*.xls*"))
    if not files:
        print("  ⚠️ 다운로드 파일 못 찾음")
        driver.close()
        driver.switch_to.window(main_window)
        return None

    latest = max(files, key=os.path.getmtime)
    print(f"  📥 다운로드: {os.path.basename(latest)}")

    driver.close()
    driver.switch_to.window(main_window)
    return latest




# ════════════════════════════════════════════════════════════
# STEP 5-3: 평가사 이름으로 검색 → 최근 평가이력(Ⅱ) 날짜 수집
# ════════════════════════════════════════════════════════════
def get_evaluator_recent_history(driver, name):
    """
    평가사관리 페이지에서 이름 검색 후 상세페이지 진입,
    평가이력(Ⅱ) 최상단 날짜 반환.

    검색창: id="searchUserName"
    이름 링크: <a href="javascript:detail(숫자);">이름</a>
    평가이력(Ⅱ): <h2 class="mt15">평가이력(Ⅱ)...</h2>
    날짜 링크: <a href="javascript:HistoryInfo(숫자)">2025-08-06&nbsp;~&nbsp;2025-08-08</a>
    """
    EVALUATOR_URL = "https://www.knab.go.kr/usr/evm/EvlprInfoInqireList.do"

    try:
        driver.get(EVALUATOR_URL)
        time.sleep(1.5)

        # 이름 검색
        search_box = driver.find_element(By.ID, "searchUserName")
        search_box.clear()
        search_box.send_keys(name)
        driver.execute_script("search();")
        time.sleep(2)

        # 이름 링크 찾기 (javascript:detail(숫자))
        links = driver.find_elements(By.XPATH, "//a[contains(@href,'detail') and text()='" + name + "']")
        if not links:
            # 이름이 정확히 일치 안할 수 있으니 포함으로도 시도
            links = driver.find_elements(By.XPATH, f"//a[contains(@href,'detail') and contains(text(),'{name}')]")
        if not links:
            return ""

        # 링크 클릭 → 새 창 또는 같은 창
        main_window = driver.current_window_handle
        links[0].click()
        time.sleep(2)

        # 새 창 열렸으면 전환
        if len(driver.window_handles) > 1:
            for handle in driver.window_handles:
                if handle != main_window:
                    driver.switch_to.window(handle)
                    break

        # 평가이력(Ⅱ) h2 찾고 그 다음 테이블의 첫 번째 날짜 링크
        recent_date = ""
        try:
            # h2 태그 중 "평가이력(Ⅱ)" 포함된 것 찾기
            h2_els = driver.find_elements(By.CSS_SELECTOR, "h2.mt15")
            target_h2 = None
            for h2 in h2_els:
                if "평가이력" in h2.text and "Ⅱ" in h2.text:
                    target_h2 = h2
                    break

            if target_h2:
                # h2 다음 테이블의 첫 번째 HistoryInfo 링크
                date_link = driver.find_element(
                    By.XPATH,
                    "//h2[contains(text(),'평가이력') and contains(text(),'Ⅱ')]"
                    "/following::a[contains(@href,'HistoryInfo')][1]"
                )
                recent_date = date_link.text.replace(" ", " ").strip()
        except Exception:
            pass

        # 창 정리
        if len(driver.window_handles) > 1:
            driver.close()
            driver.switch_to.window(main_window)

        return recent_date

    except Exception as e:
        print(f"    ⚠️ {name} 이력 조회 오류: {e}")
        return ""

# ════════════════════════════════════════════════════════════
# STEP 5-2: 평가사관리 → 대분류/중분류별 엑셀 다운로드
# ════════════════════════════════════════════════════════════
def download_evaluator_excels(driver, sheet_names, range_file_name=''):
    """
    평가사관리 페이지에서 시트명 목록(예: ["03.014", "02.017"])을 받아
    각 대분류/중분류 드롭다운 설정 → 검색 → 엑셀 다운로드.

    드롭다운:
      accreditClss: 인정분류 (항상 "02"=시험)
      searchBigId:  대분류 코드 (onchange로 중분류 동적 로드됨)
      searchMiddleId: 중분류 코드
    엑셀 버튼: onclick="download()"

    반환: {"03.014": "C:/Users/user/Downloads/xxx.xls", ...}
    """
    EVALUATOR_URL = "https://www.knab.go.kr/usr/evm/EvlprInfoInqireList.do"
    downloaded = {}

    for sheet_name in sorted(sheet_names):
        parts = sheet_name.split(".")
        if len(parts) < 2:
            continue
        대분류코드 = parts[0]  # 예: "03"
        중분류코드 = parts[1]  # 예: "014"


        print(f"  🔍 평가사 검색: {sheet_name}")

        try:
            driver.get(EVALUATOR_URL)
            time.sleep(2)

            # 1. 인정분류 선택 (파일명으로 자동 구분)
            # TestingRangeList → 시험(02), CalibarationRangeList → 교정(01)
            if "Calibar" in (range_file_name or "") or "Calib" in (range_file_name or ""):
                accredit_cls = "01"  # 교정
            else:
                accredit_cls = "02"  # 시험 (기본값)
            Select(driver.find_element(By.ID, "accreditClss")).select_by_value(accredit_cls)
            time.sleep(1)

            # 2. 대분류 선택 (onchange로 중분류 목록 동적 로드)
            Select(driver.find_element(By.ID, "searchBigId")).select_by_value(대분류코드)
            time.sleep(1.5)  # 중분류 목록 로드 대기

            # 3. 중분류 선택
            Select(driver.find_element(By.ID, "searchMiddleId")).select_by_value(중분류코드)
            time.sleep(1)

            # 소분류 선택 없음 (교정/시험 모두 대분류.중분류까지만)

            # 5. 검색
            driver.execute_script("search();")
            time.sleep(2)

            # 5. 다운로드 전 기존 파일 목록 스냅샷
            before = set(glob.glob(os.path.join(DOWNLOAD_DIR, "*.xls*")))

            # 6. 엑셀 다운로드
            driver.execute_script("download();")
            time.sleep(4)

            # 7. 새로 생긴 파일 찾기
            after = set(glob.glob(os.path.join(DOWNLOAD_DIR, "*.xls*")))
            new_files = after - before
            if new_files:
                latest = max(new_files, key=os.path.getmtime)
                downloaded[sheet_name] = latest
                print(f"    ✅ 다운로드: {os.path.basename(latest)}")
            else:
                print(f"    ⚠️ 다운로드 파일 없음: {sheet_name}")

        except Exception as e:
            print(f"    ⚠️ 평가사 검색 오류 ({sheet_name}): {e}")

    return downloaded

# ════════════════════════════════════════════════════════════
# STEP 6: 기관정보 팝업 → 기관명, 담당자정보, 사업장주소 수집
# ════════════════════════════════════════════════════════════
def get_agency_info(driver):
    """
    기관정보보기 팝업에서 기관명, 담당자 정보, 사업장 주소 수집.
    버튼: <input type="button" value="기관정보보기" onclick="PopupCompInfo('companyNo','01','02');" id="CompInfoPopUp">

    팝업 HTML 특이사항:
    1. 사업장 섹션 감지:
       <th rowspan="6">사<br>업<br>장</th> 형태로 세로 병합
       → th.text = "사\n업\n장" → replace("\n","") 후 "사업장" 감지
       → in_company_section 플래그를 True로 설정

    2. 주소 수집:
       <th>주소</th><td colspan="3">광주광역시 북구 첨단과기로 333</td>
       → th.find_element(XPATH, "following-sibling::td[1]")로 td 가져옴
       → 법인/사업장 두 곳에 "주소" th가 있으므로
         in_company_section=True 이후 첫 번째 주소만 수집

    3. 담당자 정보는 팝업 하단에 위치 (스크롤 없이도 DOM에 있음)
    """
    result = {"기관명": "", "담당자명": "", "담당자연락처": "", "담당자이메일": "", "주소": ""}
    main_window = driver.current_window_handle

    clicked = click_btn(
        driver,
        "//input[contains(@value,'기관정보보기')]",
        "//a[contains(text(),'기관정보보기')]"
    )
    if not clicked:
        print("  ⚠️ 기관정보보기 버튼 없음")
        return result

    switch_to_popup(driver, main_window)
    time.sleep(2)

    try:
        rows = driver.find_elements(By.CSS_SELECTOR, "table tr")
        in_company_section = False  # 사업장 섹션 진입 여부 플래그

        for row in rows:
            ths = row.find_elements(By.TAG_NAME, "th")
            tds = row.find_elements(By.TAG_NAME, "td")

            # 사업장 rowspan th 감지
            for th in ths:
                th_text = th.text.replace("\n", "").replace(" ", "")
                if "사업장" in th_text:
                    in_company_section = True

            # th/td가 같은 행에 있는 경우 (기관명, 주소 등)
            for th in ths:
                label = th.text.replace("\n", "").strip()
                try:
                    td = th.find_element(By.XPATH, "following-sibling::td[1]")
                    val = td.text.strip()
                except Exception:
                    val = ""

                if "기관명" in label and not result["기관명"] and "영문" not in label:
                    result["기관명"] = val
                elif label == "주소" and in_company_section and not result["주소"]:
                    result["주소"] = val

            # 담당자명/연락처/이메일: 모두 같은 tr 안에 th + td 함께 있음
            # <tr><th>담당자명</th><td>류제형</td></tr>
            for th in ths:
                label = th.text.replace("\n", "").strip()
                try:
                    val = th.find_element(By.XPATH, "following-sibling::td[1]").text.strip()
                except Exception:
                    val = ""
                if "담당자명" in label and not result["담당자명"]:
                    result["담당자명"] = val
                elif "담당자연락처" in label:
                    result["담당자연락처"] = val
                elif "담당자이메일" in label:
                    result["담당자이메일"] = val

    except Exception as e:
        print(f"  ⚠️ 기관정보 파싱 오류: {e}")

    driver.close()
    driver.switch_to.window(main_window)
    return result


# ════════════════════════════════════════════════════════════
# STEP 7: 엑셀 템플릿 복사 후 데이터 입력 저장
# ════════════════════════════════════════════════════════════
def fill_excel(data):
    """
    template.xlsx를 복사하여 '1안' 시트에 데이터 입력 후 저장.

    인정분야 데이터:
    - xlrd로 xls 읽기 (openpyxl은 .xls 미지원, xlrd 사용)
    - 1행(인덱스0)은 헤더 → 2행(인덱스1)부터 '1안' 시트 19행에 붙여넣기
    - 대분류코드(B열=인덱스1) + 중분류코드(D열=인덱스3) 조합으로 새 시트 생성
      xls 숫자는 float으로 읽힘 → int 변환 후 zfill 처리
      예) 1.0 → "01", 17.0 → "017" → 시트명 "01.017"

    파일명: {기관명}_{평가종류}_평가계획서(안).xlsx
    저장위치: OUTPUT_DIR (ahnhyoseung/)
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    기관명   = data.get("기관명", "unknown")
    평가종류 = data.get("평가종류", "")
    # 파일명 특수문자 제거
    safe_name = re.sub(r'[\\/*?:"<>|]', "_", 기관명)
    out_path  = os.path.join(OUTPUT_DIR, f"{safe_name}_{평가종류}_평가계획서(안).xlsx")

    # 템플릿 복사 후 열기
    shutil.copy2(TEMPLATE_FILE, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb["1안"]  # '1안' 시트에 입력

    # 기본 정보
    # 병합 셀 안전 쓰기 함수
    def safe_write(ws, cell_addr, value):
        from openpyxl.utils import coordinate_to_tuple
        row, col = coordinate_to_tuple(cell_addr)
        target_merge = None
        for merge in list(ws.merged_cells.ranges):
            if merge.min_row <= row <= merge.max_row and merge.min_col <= col <= merge.max_col:
                target_merge = str(merge)
                break
        if target_merge:
            ws.unmerge_cells(target_merge)
            ws[cell_addr] = value
            ws.merge_cells(target_merge)
        else:
            ws[cell_addr] = value

    ws["B3"] = 기관명
    ws["D3"] = 평가종류

    # 내부심사자 (B4부터 옆으로 계속)
    심사자 = data.get("내부심사", [])
    cols_4 = ["B","C","D","E","F","G","H","I","J","K"]
    for i, 이름 in enumerate(심사자):
        if i < len(cols_4):
            safe_write(ws, f"{cols_4[i]}4", 이름)

    # 컨설팅 담당자 (B5부터 옆으로 계속)
    컨설팅 = data.get("컨설팅", [])
    for i, 이름 in enumerate(컨설팅):
        if i < len(cols_4):
            safe_write(ws, f"{cols_4[i]}5", 이름)

    # 직전 평가반 (B6~H6, 최대 7명)
    cols  = ["B", "C", "D", "E", "F", "G", "H"]
    평가반 = data.get("직전평가반", [])
    for i, col in enumerate(cols):
        ws[f"{col}6"] = 평가반[i] if i < len(평가반) else ""

    # 평가일
    ws["B7"] = data.get("평가시작일", "")
    ws["C7"] = data.get("평가종료일", "")



    safe_write(ws, "B14", data.get("담당자명", ""))
    safe_write(ws, "B15", data.get("담당자연락처", ""))
    safe_write(ws, "B16", data.get("담당자이메일", ""))
    safe_write(ws, "F14", data.get("주소", ""))

    # 평가반구성 (10행부터): B=역할, C=소속, D=성명, E=연락처
    # 기존 내용 먼저 클리어
    평가반구성 = data.get("평가반구성", [])
    for i, 멤버 in enumerate(평가반구성):
        row = 10 + i
        safe_write(ws, f"B{row}", 멤버.get("소속", ""))
        safe_write(ws, f"C{row}", 멤버.get("성명", ""))
        safe_write(ws, f"D{row}", 멤버.get("연락처", ""))
        safe_write(ws, f"E{row}", 멤버.get("이메일", ""))

    # 인정분야 데이터 붙여넣기 + 대분류.중분류 시트 생성
    range_file = data.get("인정분야파일", None)
    if range_file and os.path.exists(range_file):
        try:
            # xlrd로 .xls 파일 읽기 (openpyxl은 .xls 미지원)
            wb_range = xlrd.open_workbook(range_file)
            ws_range = wb_range.sheet_by_index(0)

            # 기존 데이터 클리어: 19행부터 끝까지 비우기 (이전 데이터 잔존 방지)
            for r in ws.iter_rows(min_row=19, max_row=ws.max_row):
                for cell in r:
                    cell.value = None

            # '1안' 시트 19행부터 붙여넣기 (xls 2행=인덱스1부터, 1행은 헤더)
            start_row = 19
            for row_idx in range(1, ws_range.nrows):
                for col_idx in range(ws_range.ncols):
                    val = ws_range.cell_value(row_idx, col_idx)
                    ws.cell(row=start_row, column=col_idx + 1, value=val)
                start_row += 1
            print(f"  📋 인정분야 {start_row - 19}행 복사 완료")

            # 대분류코드(B열=인덱스1) + 중분류코드(D열=인덱스3) 조합으로 시트 생성
            sheet_names = set()
            # 교정 여부 판단 (파일명으로 구분)
            is_calibration = range_file and "Calibar" in os.path.basename(range_file)

            for row_idx in range(1, ws_range.nrows):
                대분류_val = ws_range.cell_value(row_idx, 1)  # B열 = 대분류코드
                중분류_val = ws_range.cell_value(row_idx, 3)  # D열 = 중분류코드
                대분류 = str(int(float(대분류_val))).zfill(2) if 대분류_val else ""
                중분류 = str(int(float(중분류_val))).zfill(3) if 중분류_val else ""
                if 대분류 and 중분류:
                    if is_calibration:
                        소분류_val = ws_range.cell_value(row_idx, 5)  # F열 = 소분류코드
                        소분류 = str(int(float(소분류_val))).zfill(3) if 소분류_val else ""
                        if 소분류:
                            sheet_names.add(f"{대분류}.{중분류}.{소분류}")
                    else:
                        sheet_names.add(f"{대분류}.{중분류}")

            # 평가사관리에서 각 시트별 엑셀 다운로드 후 시트에 붙여넣기
            evaluator_files = data.get("평가사파일들", {})
            sheet_names_sorted = sorted(sheet_names)
            for i, sheet_name in enumerate(sheet_names_sorted):
                # 1. 시트 생성
                if sheet_name not in wb.sheetnames:
                    if i == 0:
                        wb["Sheet1"].title = sheet_name
                        print(f"  📄 시트 이름 변경: Sheet1 → {sheet_name}")
                    else:
                        # 서식 문제 방지: copy_worksheet 대신 빈 시트 생성 후 1행 헤더만 복사
                        source = wb[sheet_names_sorted[0]]
                        new_ws = wb.create_sheet(title=sheet_name)
                        # 1행 헤더만 복사 (서식 없이 값만)
                        for col_idx, cell in enumerate(source[1], 1):
                            new_ws.cell(row=1, column=col_idx, value=cell.value)
                        print(f"  📄 시트 생성: {sheet_name}")

                # 2. 해당 시트에 맞는 평가사 파일 붙여넣기
                eval_file = evaluator_files.get(sheet_name)
                if not eval_file or not os.path.exists(eval_file):
                    continue
                try:
                    ws_eval = wb[sheet_name]
                    wb_eval = xlrd.open_workbook(eval_file)
                    ws_src = wb_eval.sheet_by_index(0)

                    # 사이트 엑셀 → 템플릿 열 매핑
                    # C(3)→A(1), D(4)→B(2), G(7)→C(3), P(16)→D(4), K(11)→E(5)
                    col_map = {3: 1, 4: 2, 7: 3, 16: 4, 11: 5}

                    data_rows = 0
                    for r_idx in range(1, ws_src.nrows):
                        row_has_data = False
                        for src_col, dst_col in col_map.items():
                            val = ws_src.cell_value(r_idx, src_col - 1)  # xlrd는 0-based
                            if isinstance(val, str):
                                val = re.sub(r'<[^>]*>?', '', val).strip()
                                # 엑셀에서 사용 불가한 제어문자 제거
                                val = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', val)
                            # 평가사구분 변환: 시험:평가사보 → 평가사보
                            if dst_col == 3:
                                m = re.search(r'시험[:;,](선임평가사|평가사보|평가사)', str(val))
                                if m:
                                    val = m.group(1)
                            ws_eval.cell(row=r_idx + 1, column=dst_col, value=val)
                            if val:
                                row_has_data = True
                        if row_has_data:
                            data_rows += 1

                    if ws_eval.dimensions and ws_eval.dimensions != 'A1:A1':
                        ws_eval.auto_filter.ref = ws_eval.dimensions

                    # 특정 이름 목록에 있으면 B열(성명) 빨간색 표시
                    from openpyxl.styles import PatternFill, Font
                    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    red_font = Font(color="FF0000", bold=True)
                    red_names = set(["강민철","고우성","김경식","김경희","김남진","김진규","김진수","김태성","김하동","김학영","김희수","박갑동","박경환","박덕우","박상철","송지훈","심규창","안완식","안장혁","엄석원","오상만","유덕룡","유찬주","육선우","이규배","이명수","이상문","이석기","이승덕","이영규","이종대","이창수","장성우","장태연","주정우","조대흥","차영섭","최성진","탁계성","황병옥","황수환"])
                    for r_idx in range(2, ws_eval.max_row + 1):
                        name_cell = ws_eval.cell(row=r_idx, column=2)  # B열=성명
                        cell_name = str(name_cell.value).strip() if name_cell.value else ""
                        if cell_name == "고우성":
                            print(f"  🔍 고우성 발견 row={r_idx}, repr={repr(cell_name)}, in_red={cell_name in red_names}")
                        if cell_name and cell_name in red_names:
                            name_cell.font = red_font

                    print(f"  📋 시트 [{sheet_name}] {data_rows}행 복사 완료")
                except Exception as e:
                    print(f"  ⚠️ 시트 [{sheet_name}] 복사 오류: {e}")
                    import traceback
                    traceback.print_exc()

        except Exception as e:
            print(f"  ⚠️ 인정분야 복사 오류: {e}")

    wb.save(out_path)
    print(f"  💾 저장: {out_path}")
    return out_path


# ════════════════════════════════════════════════════════════
# 메인
# ════════════════════════════════════════════════════════════
def main():
    print("=" * 60)
    print("  KOLAS 자동화 프로그램 v2.0")
    print("=" * 60)

    if not os.path.exists(TEMPLATE_FILE):
        print(f"❌ 템플릿 파일 없음: {TEMPLATE_FILE}")
        return

    driver = setup_driver()
    results = []

    try:
        login(driver)
        items = collect_list(driver)

        if not items:
            print("❌ 수집된 항목 없음.")
            return

        for idx, item in enumerate(items, 1):
            접수번호 = item["접수번호"]
            print(f"\n[{idx}/{len(items)}] 접수번호: {접수번호}")

            try:
                accredit = item["accreditNo"]

                # 평가종류: 상세페이지 본문 "신청분류" 셀에서 수집
                # (신규/정기검사/재평가/확대 중 하나)
                평가종류 = ""
                try:
                    rows = driver.find_elements(By.CSS_SELECTOR, "table tr")
                    for row in rows:
                        ths = row.find_elements(By.TAG_NAME, "th")
                        tds = row.find_elements(By.TAG_NAME, "td")
                        for i, th in enumerate(ths):
                            if "신청분류" in th.text.strip():
                                평가종류 = tds[i].text.strip() if i < len(tds) else ""
                                break
                        if 평가종류:
                            break
                except Exception:
                    pass

                # 평가반구성 및 업무분장: 상세페이지 본문에서 직접 수집
                평가반구성 = []
                try:
                    level_els  = driver.find_elements(By.CSS_SELECTOR, "select[name='assessorLevel']")
                    name_els   = driver.find_elements(By.CSS_SELECTOR, "input[name='assessorNameArr']")
                    office_els = driver.find_elements(By.CSS_SELECTOR, "input[name='assessorOfficeArr']")
                    phone_els  = driver.find_elements(By.CSS_SELECTOR, "input[name='assessorPhoneArr']")
                    email_els  = driver.find_elements(By.CSS_SELECTOR, "input[name='assessorEmailArr']")
                    for i in range(len(name_els)):
                        try:
                            역할 = level_els[i].find_element(By.CSS_SELECTOR, "option[selected]").text.strip() if i < len(level_els) else ""
                        except Exception:
                            역할 = ""
                        성명 = name_els[i].get_attribute("value").strip()
                        소속 = office_els[i].get_attribute("value").strip() if i < len(office_els) else ""
                        연락처 = phone_els[i].get_attribute("value").strip() if i < len(phone_els) else ""
                        이메일 = email_els[i].get_attribute("value").strip() if i < len(email_els) else ""
                        if 성명:
                            평가반구성.append({"역할": 역할, "소속": 소속, "성명": 성명, "연락처": 연락처, "이메일": 이메일})
                    print(f"  👥 평가반구성: {len(평가반구성)}명")
                except Exception as e:
                    print(f"  ⚠️ 평가반구성 수집 오류: {e}")

                # 팝업 수집 순서 (이 순서 중요!)
                # 상세 페이지는 POST 방식이라 팝업 닫고 URL 재접근 불가
                # 따라서 모든 팝업을 한 번의 상세페이지 진입 상태에서 처리

                # 1. 기관정보보기 팝업 → 기관명, 담당자정보, 사업장주소
                agency = get_agency_info(driver)

                # 2. 컨설팅정보보기 팝업 → 내부심사자, 컨설팅담당자
                consulting = get_consulting_info(driver)

                # 3. 인정분야보기 팝업 → 엑셀 다운로드
                range_file = get_accredit_range(driver)

                # 4. 기관이력보기 팝업 → 직전 평가반 (마지막 수행)
                evaluators = get_history_evaluators(driver, accredit)

                # 5. 평가사관리 페이지에서 대분류.중분류별 평가사 엑셀 다운로드
                # (인정분야 xls에서 추출한 sheet_names 사용)
                평가사파일들 = {}
                if range_file and os.path.exists(range_file):
                    try:
                        wb_tmp = xlrd.open_workbook(range_file)
                        ws_tmp = wb_tmp.sheet_by_index(0)
                        sheet_names_tmp = set()
                        for r in range(1, ws_tmp.nrows):
                            대v = ws_tmp.cell_value(r, 1)
                            중v = ws_tmp.cell_value(r, 3)
                            대 = str(int(float(대v))).zfill(2) if 대v else ""
                            중 = str(int(float(중v))).zfill(3) if 중v else ""
                            if 대 and 중:
                                sheet_names_tmp.add(f"{대}.{중}")
                        평가사파일들 = download_evaluator_excels(driver, sheet_names_tmp, os.path.basename(range_file) if range_file else '')
                    except Exception as e:
                        print(f"  ⚠️ 평가사 다운로드 준비 오류: {e}")

                data = {
                    "기관명":       agency.get("기관명", ""),
                    "평가종류":     평가종류,
                    "내부심사":     consulting.get("내부심사", []),
                    "컨설팅":       consulting.get("컨설팅", []),
                    "직전평가반":   evaluators,
                    "평가시작일":   "",   # 현재 미사용 (추후 구현 가능)
                    "평가종료일":   "",   # 현재 미사용 (추후 구현 가능)
                    "담당자명":     agency.get("담당자명", ""),
                    "담당자연락처": agency.get("담당자연락처", ""),
                    "담당자이메일": agency.get("담당자이메일", ""),
                    "주소":         agency.get("주소", ""),
                    "인정분야파일": range_file,
                    "평가사파일들":  평가사파일들,
                    "평가반구성":   평가반구성,
                }

                print(f"  기관명:     {data['기관명']}")
                print(f"  평가종류:   {data['평가종류']}")
                print(f"  내부심사:   {data['내부심사']}")
                print(f"  컨설팅:     {data['컨설팅']}")
                print(f"  직전평가반: {data['직전평가반']}")
                print(f"  담당자:     {data['담당자명']} / {data['담당자연락처']}")

                out = fill_excel(data)
                results.append({"접수번호": 접수번호, "기관명": data["기관명"], "파일": out})

            except Exception as e:
                print(f"  ❌ 처리 중 오류: {e}")
                import traceback
                traceback.print_exc()
                # 오류 발생 시 메인 창으로 복귀 후 다음 접수번호 처리 계속
                try:
                    driver.switch_to.window(driver.window_handles[0])
                except Exception:
                    pass
                continue

    except Exception as e:
        print(f"\n❌ 치명적 오류: {e}")
        import traceback
        traceback.print_exc()

    finally:
        driver.quit()

    print("\n" + "=" * 60)
    print(f"  완료! 총 {len(results)}개 파일 → '{OUTPUT_DIR}/' 폴더")
    print("=" * 60)
    for r in results:
        print(f"  [{r['접수번호']}] {r['기관명']}")


if __name__ == "__main__":
    main()